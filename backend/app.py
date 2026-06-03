"""
FIF Analytics Dashboard — Flask Backend
Reads real Excel data and serves it via REST API.
"""
from flask import Flask, jsonify, request
from flask_cors import CORS
import openpyxl
from pathlib import Path
import re

app = Flask(__name__)
CORS(app)

BASE = Path(__file__).parent.parent  # FIF folder
FIF_FILE   = BASE / 'Financial Inclusion Fund-End of March  2026.xlsx'
BRIDGE_FILE = BASE / 'Bridge End Month of March 2026.xlsx'
TELCO_FILE  = BASE / 'Financial Inlcusion Fund-End of March 2026 {Airtel  Telkom } - Individual (1).xlsx'

MONTHS = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def safe(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def month_label(yyyymm):
    s = str(int(yyyymm))
    return f"{MONTHS[int(s[4:6])]} {s[:4]}"


def telco_to_yyyymm(val):
    """Convert 'Nov-2022' or 'Sept-2023' → '202211'"""
    mo = {'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
          'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
          'sep': '09', 'sept': '09', 'oct': '10', 'nov': '11', 'dec': '12'}
    m = re.match(r'([A-Za-z]+)[-\s](\d{4})', str(val).strip())
    if m:
        return f"{m.group(2)}{mo.get(m.group(1).lower(), '00')}"
    return None


# ─────────────────────────────────────────────
# Readers
# ─────────────────────────────────────────────

def read_monthly(filepath, sheet='Monthly Summary', svgs_factor=1_000_000):
    """
    Read a standard Monthly Summary sheet.
    svgs_factor: FIF stores mandatory savings in Mn (×1M), Bridge stores raw KES (×1).
    Due / Outstanding are stored in Mn for both → always ×1M.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        mnth = row[1]
        if not (mnth and isinstance(mnth, (int, float)) and int(mnth) > 200000):
            continue
        rows.append({
            'month':        str(int(mnth)),
            'label':        month_label(mnth),
            'custBase':     safe(row[2]),
            'mandatorySvgs': safe(row[3]) * svgs_factor,
            'disbVol':      safe(row[4]),
            'disbVal':      safe(row[5]),
            'disbCust':     safe(row[6]),
            'avgTicket':    safe(row[7]),
            'repayVol':     safe(row[8]),
            'repayVal':     safe(row[9]),
            'repayCust':    safe(row[10]),
            'due':          safe(row[11]) * 1_000_000,
            'outstanding':  safe(row[12]) * 1_000_000,
            'repayRate':    safe(row[13]),
        })
    wb.close()
    return rows


def read_telco_monthly():
    """Read Airtel + Telkom monthly sheets and merge by month."""
    wb = openpyxl.load_workbook(TELCO_FILE, read_only=True, data_only=True)

    def read_sheet(name):
        ws = wb[name]
        out = {}
        for row in ws.iter_rows(min_row=5, values_only=True):
            raw = row[1]
            if not raw or str(raw).strip() in ('', 'ID_MNTH', 'Individuals'):
                continue
            m = telco_to_yyyymm(raw)
            if not m:
                continue
            out[m] = {
                'disbVol':  safe(row[4]),
                'disbVal':  safe(row[5]),
                'avgTicket': safe(row[7]),
                'repayVal': safe(row[9]),
            }
        return out

    airtel = read_sheet('Monthly Summary (Airtel)')
    telkom = read_sheet('Monthly Summary (Telkom)')
    wb.close()

    all_months = sorted(set(list(airtel) + list(telkom)))
    result = []
    for m in all_months:
        a = airtel.get(m, {})
        t = telkom.get(m, {})
        result.append({
            'month':       m,
            'label':       month_label(int(m)),
            'airtelVal':   a.get('disbVal', 0),
            'telkomVal':   t.get('disbVal', 0),
            'airtelVol':   a.get('disbVol', 0),
            'telkomVol':   t.get('disbVol', 0),
            'airtelTicket': a.get('avgTicket', 0),
            'telkomTicket': t.get('avgTicket', 0),
        })
    return result


def read_daily_aggregates():
    """
    Read FIF Daily Summary(Individuals) — last row per month is the monthly cumulative total.
    Returns interest (accrued/paid) and savings (mandatory/voluntary/withdrawn) by month.
    """
    wb = openpyxl.load_workbook(FIF_FILE, read_only=True, data_only=True)
    ws = wb['Daily Summary(Individuals)']
    by_month = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        mnth = row[1]
        if not mnth or str(mnth).strip() in ('', 'ID_MNTH'):
            continue
        by_month[str(mnth).strip()] = {
            'accrued':      safe(row[14]),
            'paid':         safe(row[15]),
            'mandatoryVol': safe(row[16]),
            'mandatoryVal': safe(row[17]),
            'voluntaryVol': safe(row[18]),
            'voluntaryVal': safe(row[19]),
            'voluntarySubs': safe(row[20]),
            'withdrawnVol': safe(row[21]),
            'withdrawnVal': safe(row[22]),
            'withdrawnSubs': safe(row[23]),
        }
    wb.close()
    return by_month


# ─────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/api/monthly')
def monthly():
    product = request.args.get('product', 'FIF').upper()
    if product == 'BRIDGE':
        # Bridge stores mandatory savings in raw KES (svgs_factor=1)
        return jsonify(read_monthly(BRIDGE_FILE, svgs_factor=1))
    # FIF stores mandatory savings in Mn (svgs_factor=1_000_000)
    return jsonify(read_monthly(FIF_FILE, svgs_factor=1_000_000))


@app.route('/api/telco')
def telco():
    return jsonify(read_telco_monthly())


@app.route('/api/interest')
def interest():
    by_month = read_daily_aggregates()
    rows = []
    for m in sorted(by_month):
        d = by_month[m]
        rows.append({
            'month':   m,
            'label':   month_label(int(m)),
            'accrued': d['accrued'],
            'paid':    d['paid'],
        })
    return jsonify(rows)


@app.route('/api/savings')
def savings():
    by_month = read_daily_aggregates()
    rows = []
    for m in sorted(by_month):
        d = by_month[m]
        rows.append({
            'month':        m,
            'label':        month_label(int(m)),
            'mandatoryVal': d['mandatoryVal'],
            'mandatoryVol': d['mandatoryVol'],
            'voluntaryVal': d['voluntaryVal'],
            'voluntaryVol': d['voluntarySubs'],   # using subscriber count as "volume"
            'withdrawnVal': d['withdrawnVal'],
            'withdrawnVol': d['withdrawnSubs'],
        })
    return jsonify(rows)


def append_to_excel(filepath, record, svgs_factor=1_000_000):
    """
    Append a new row to the Monthly Summary sheet of an Excel file.
    Form values:  mandatorySvgs in Mn, due in Mn, outstanding in Mn, disbVal/repayVal in raw KES.
    Excel stores: mandatorySvgs in Mn for FIF (÷1M already), raw KES for Bridge (×1M).
                  due and outstanding always in Mn.
    svgs_factor: 1_000_000 = FIF (store as Mn), 1 = Bridge (store raw KES → multiply back)
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb['Monthly Summary']

    # Find first empty row after data (scan column B for last filled row)
    last_row = ws.max_row
    while last_row > 4 and ws.cell(last_row, 2).value in (None, ''):
        last_row -= 1
    new_row = last_row + 1

    month_int = int(str(record['month']).strip())
    # For Bridge, mandatory savings is stored raw KES in Excel; for FIF it's stored as Mn
    svgs_excel = record['mandatorySvgs'] / svgs_factor  # Mn → Mn (FIF) or Mn → raw KES (Bridge ×1M ÷1=Mn? No)
    # Actually: user enters Mn. FIF Excel stores Mn → write as-is (factor=1M means divide by 1M? No…)
    # read_monthly multiplies row[3] * svgs_factor to get KES.
    # So Excel value = KES_value / svgs_factor
    # User entered Mn → KES = user_Mn * 1_000_000 → Excel = (user_Mn * 1_000_000) / svgs_factor
    svgs_excel = (record['mandatorySvgs'] * 1_000_000) / svgs_factor

    ws.cell(new_row, 2).value  = month_int
    ws.cell(new_row, 3).value  = record['custBase']
    ws.cell(new_row, 4).value  = svgs_excel
    ws.cell(new_row, 5).value  = record['disbVol']
    ws.cell(new_row, 6).value  = record['disbVal']
    ws.cell(new_row, 7).value  = record['disbCust']
    ws.cell(new_row, 8).value  = record['avgTicket']
    ws.cell(new_row, 9).value  = record['repayVol']
    ws.cell(new_row, 10).value = record['repayVal']
    ws.cell(new_row, 11).value = record['repayCust']
    ws.cell(new_row, 12).value = record['due']         # already Mn
    ws.cell(new_row, 13).value = record['outstanding']  # already Mn
    ws.cell(new_row, 14).value = record['repayRate']

    wb.save(filepath)
    wb.close()


@app.route('/api/records', methods=['POST'])
def create_record():
    """Accept a new monthly record and write it to the appropriate Excel file."""
    body = request.get_json(force=True, silent=True)
    print(f'[POST /api/records] Received: {body}')  # Log for debugging
    
    if not body:
        return jsonify({'error': 'Invalid JSON'}), 400
    if not body.get('month') or not body.get('product'):
        return jsonify({'error': 'month and product are required'}), 400

    record = {
        'month':       str(body.get('month', '')).strip(),
        'product':     body.get('product', 'FIF').upper(),
        'segment':     body.get('segment', 'Individuals'),
        'telco':       body.get('telco', 'All'),
        'custBase':    float(body.get('custBase') or 0),
        'mandatorySvgs': float(body.get('mandatorySvgs') or 0),
        'disbVol':     float(body.get('disbVol') or 0),
        'disbVal':     float(body.get('disbVal') or 0),
        'disbCust':    float(body.get('disbCust') or 0),
        'avgTicket':   float(body.get('avgTicket') or 0),
        'repayVol':    float(body.get('repayVol') or 0),
        'repayVal':    float(body.get('repayVal') or 0),
        'repayCust':   float(body.get('repayCust') or 0),
        'repayRate':   float(body.get('repayRate') or 0),
        'due':         float(body.get('due') or 0),
        'outstanding': float(body.get('outstanding') or 0),
    }

    try:
        if record['product'] == 'BRIDGE':
            append_to_excel(BRIDGE_FILE, record, svgs_factor=1)
        else:
            append_to_excel(FIF_FILE, record, svgs_factor=1_000_000)
        print(f'[POST /api/records] ✓ Saved successfully: {record["month"]}, {record["product"]}')
        return jsonify({'status': 'saved', 'month': record['month']}), 201
    except Exception as e:
        print(f'[POST /api/records] ✗ Error: {str(e)}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/records/<record_id>', methods=['DELETE'])
def delete_record(record_id):
    """Placeholder — deleting from Excel is not supported via API."""
    return jsonify({'status': 'ignored', 'note': 'Deletes are local only; Excel is not modified'}), 200


if __name__ == '__main__':
    print('Starting FIF backend on http://localhost:5000')
    app.run(debug=False, port=5000)
